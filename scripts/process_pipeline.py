import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
import os
import json
from datetime import datetime, date, timedelta

def clean_and_process_data():
    raw_file = "data/raw_crm_export.csv"
    if not os.path.exists(raw_file):
        raise FileNotFoundError(f"Raw data file {raw_file} not found. Please run generate_raw_data.py first.")
        
    print("Starting ETL process...")
    
    # 1. Load data
    df = pd.read_csv(raw_file)
    initial_rows = len(df)
    
    stats = {
        "initial_rows": initial_rows,
        "duplicates_removed": 0,
        "missing_emails_filled": 0,
        "casing_fixes": 0,
        "date_format_fixes": 0,
        "date_order_fixes": 0,
        "mrr_fixes": 0
    }
    
    # 2. Handle duplicates
    # Find duplicates based on Lead_ID
    dup_mask = df.duplicated(subset=["Lead_ID"], keep='first')
    stats["duplicates_removed"] = int(dup_mask.sum())
    df = df[~dup_mask].copy()
    
    # 3. Clean Company Name (whitespace)
    df["Company_Name"] = df["Company_Name"].astype(str).str.strip()
    
    # 4. Fill missing emails
    empty_emails = df["Contact_Email"].isna() | (df["Contact_Email"] == "")
    stats["missing_emails_filled"] = int(empty_emails.sum())
    # Generate generic info email if missing
    for idx, row in df[empty_emails].iterrows():
        clean_company = str(row["Company_Name"]).lower().replace(' ', '').replace('(duplicateentry)', '')
        df.loc[idx, "Contact_Email"] = f"info@{clean_company}.com"
        
    # 5. Clean Lead Source casing
    sources_before = df["Lead_Source"].copy()
    
    def clean_source(src):
        if pd.isna(src) or str(src).strip().lower() in ["null", "", "none"]:
            return "Organic / Direct"
        s = str(src).strip().lower()
        if "google" in s:
            return "Google"
        elif "linkedin" in s:
            return "LinkedIn"
        elif "referral" in s:
            return "Referral"
        elif "outreach" in s or "cold" in s:
            return "Cold Outreach"
        elif "webinar" in s:
            return "Webinar"
        else:
            return s.title()
            
    df["Lead_Source"] = df["Lead_Source"].apply(clean_source)
    stats["casing_fixes"] = int((df["Lead_Source"] != sources_before).sum())
    
    # 6. Parse and Clean Dates
    def parse_date(date_str):
        if pd.isna(date_str) or str(date_str).strip() == "":
            return pd.NaT
        d_str = str(date_str).strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(d_str, fmt).date()
            except ValueError:
                continue
        return pd.NaT
        
    # Parse dates
    df["Lead_Date"] = df["Lead_Date"].apply(parse_date)
    df["Trial_Date"] = df["Trial_Date"].apply(parse_date)
    df["Conversion_Date"] = df["Conversion_Date"].apply(parse_date)
    
    # Check date conversions (all should parse successfully if they had text)
    # We will count date format normalization as date_format_fixes
    stats["date_format_fixes"] = int(df["Lead_Date"].notna().sum() + df["Trial_Date"].notna().sum() + df["Conversion_Date"].notna().sum())
    
    # Standardize Trial_Activated and Converted fields
    df["Trial_Activated"] = df["Trial_Activated"].fillna("No").str.strip().str.title()
    df["Converted"] = df["Converted"].fillna("No").str.strip().str.title()
    
    # Fix out-of-order dates
    date_order_issues = 0
    for idx, row in df.iterrows():
        lead_dt = row["Lead_Date"]
        trial_dt = row["Trial_Date"]
        conv_dt = row["Conversion_Date"]
        
        # Trial date before lead date
        if row["Trial_Activated"] == "Yes" and not pd.isna(trial_dt) and not pd.isna(lead_dt):
            if trial_dt < lead_dt:
                trial_dt = lead_dt + timedelta(days=2)
                df.loc[idx, "Trial_Date"] = trial_dt
                date_order_issues += 1
                
        # Conversion date before trial date or lead date
        if row["Converted"] == "Yes" and not pd.isna(conv_dt):
            base_dt = trial_dt if (row["Trial_Activated"] == "Yes" and not pd.isna(trial_dt)) else lead_dt
            if not pd.isna(base_dt) and conv_dt < base_dt:
                conv_dt = base_dt + timedelta(days=10)
                df.loc[idx, "Conversion_Date"] = conv_dt
                date_order_issues += 1
                
    stats["date_order_fixes"] = date_order_issues
    
    # 7. Clean and Standardize MRR
    mrr_issues = 0
    df["MRR"] = pd.to_numeric(df["MRR"], errors='coerce')
    
    for idx, row in df.iterrows():
        converted = row["Converted"] == "Yes"
        mrr = row["MRR"]
        seg = row["Segment"]
        
        # Determine standard median MRR by segment
        median_mrr = 200 if seg == "SMB" else (1000 if seg == "Mid-Market" else 5000)
        
        if converted:
            if pd.isna(mrr) or mrr <= 0 or mrr > 20000:
                # Outlier or invalid conversion MRR
                df.loc[idx, "MRR"] = median_mrr
                mrr_issues += 1
        else:
            if not pd.isna(mrr) and mrr != 0:
                df.loc[idx, "MRR"] = np.nan # Unconverted shouldn't have MRR in our flat tracker
                
    stats["mrr_fixes"] = mrr_issues
    stats["final_clean_rows"] = len(df)
    
    # Save stats to json for the web dashboard to consume
    os.makedirs("reports", exist_ok=True)
    with open("reports/cleaning_stats.json", "w") as f:
        json.dump(stats, f, indent=4)
        
    print(f"ETL completed. Removed {stats['duplicates_removed']} duplicates, resolved {stats['mrr_fixes']} MRR anomalies.")
    
    return df, stats

def create_excel_tracker(df, stats):
    print("Generating styled Excel workbook...")
    wb = openpyxl.Workbook()
    
    # Define Styles
    font_name = "Segoe UI"
    font_title = Font(name=font_name, size=16, bold=True, color="FFFFFF")
    font_section = Font(name=font_name, size=12, bold=True, color="2B3E50")
    font_header = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    font_data = Font(name=font_name, size=11)
    font_data_bold = Font(name=font_name, size=11, bold=True)
    font_kpi_num = Font(name=font_name, size=20, bold=True, color="1F4E78")
    font_kpi_label = Font(name=font_name, size=9, color="595959")
    
    fill_navy = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_slate = PatternFill(start_color="2B3E50", end_color="2B3E50", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")
    fill_kpi = PatternFill(start_color="EBF1F5", end_color="EBF1F5", fill_type="solid")
    fill_total = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    border_thin_side = Side(style='thin', color='D3D3D3')
    border_double_bottom = Side(style='double', color='000000')
    border_thin_top = Side(style='thin', color='000000')
    
    border_grid = Border(left=border_thin_side, right=border_thin_side, top=border_thin_side, bottom=border_thin_side)
    border_kpi = Border(left=border_thin_side, right=border_thin_side, top=border_thin_side, bottom=border_thin_side)
    border_total = Border(top=border_thin_top, bottom=border_double_bottom)
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    # -------------------------------------------------------------
    # TAB 1: EXECUTIVE DASHBOARD
    # -------------------------------------------------------------
    ws_dash = wb.active
    ws_dash.title = "Executive Dashboard"
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_dash.merge_cells("A1:H2")
    title_cell = ws_dash["A1"]
    title_cell.value = " B2B SaaS Revenue Pipeline Dashboard (12-Week RevOps Summary)"
    title_cell.font = font_title
    title_cell.fill = fill_navy
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Add KPI Cards
    # KPI 1: Total Leads
    ws_dash.merge_cells("A4:B4")
    ws_dash["A4"] = "TOTAL LEADS"
    ws_dash["A4"].font = font_kpi_label
    ws_dash["A4"].alignment = align_center
    ws_dash["A4"].fill = fill_kpi
    
    ws_dash.merge_cells("A5:B5")
    ws_dash["A5"] = "=SUM('Weekly Metrics'!B6:B18)"
    ws_dash["A5"].font = font_kpi_num
    ws_dash["A5"].alignment = align_center
    ws_dash["A5"].fill = fill_kpi
    
    # KPI 2: Total Conversions
    ws_dash.merge_cells("C4:D4")
    ws_dash["C4"] = "TOTAL CONVERSIONS"
    ws_dash["C4"].font = font_kpi_label
    ws_dash["C4"].alignment = align_center
    ws_dash["C4"].fill = fill_kpi
    
    ws_dash.merge_cells("C5:D5")
    ws_dash["C5"] = "=SUM('Weekly Metrics'!D6:D18)"
    ws_dash["C5"].font = font_kpi_num
    ws_dash["C5"].alignment = align_center
    ws_dash["C5"].fill = fill_kpi
    
    # KPI 3: Total MRR Added
    ws_dash.merge_cells("E4:F4")
    ws_dash["E4"] = "NEW MRR ADDED"
    ws_dash["E4"].font = font_kpi_label
    ws_dash["E4"].alignment = align_center
    ws_dash["E4"].fill = fill_kpi
    
    ws_dash.merge_cells("E5:F5")
    ws_dash["E5"] = "=SUM('Weekly Metrics'!G6:G18)"
    ws_dash["E5"].font = font_kpi_num
    ws_dash["E5"].alignment = align_center
    ws_dash["E5"].fill = fill_kpi
    ws_dash["E5"].number_format = "$#,##0"
    
    # KPI 4: Funnel Conversion Rate
    ws_dash.merge_cells("G4:H4")
    ws_dash["G4"] = "FUNNEL CONVERSION"
    ws_dash["G4"].font = font_kpi_label
    ws_dash["G4"].alignment = align_center
    ws_dash["G4"].fill = fill_kpi
    
    ws_dash.merge_cells("G5:H5")
    ws_dash["G5"] = "=E5/A5" # Total Conversions / Total Leads
    ws_dash["G5"].font = font_kpi_num
    ws_dash["G5"].alignment = align_center
    ws_dash["G5"].fill = fill_kpi
    ws_dash["G5"].number_format = "0.0%"
    
    # Apply borders to KPI cards
    for row in range(4, 6):
        for col in range(1, 9):
            ws_dash.cell(row=row, column=col).border = border_kpi
            
    # Add Section Header
    ws_dash["A7"] = "Key Performance Trends"
    ws_dash["A7"].font = font_section
    
    # Add a brief text overview
    ws_dash.merge_cells("A8:H8")
    ws_dash["A8"] = "This tracker maps a simulated B2B SaaS funnel showing marketing drop-offs, onboarding speed, and MRR contribution."
    ws_dash["A8"].font = Font(name=font_name, size=10, italic=True)
    
    # Data Cleansing Report on Dashboard
    ws_dash["A10"] = "RevOps ETL Data Audit Report"
    ws_dash["A10"].font = font_section
    
    ws_dash["A11"] = "Metric"
    ws_dash["A11"].font = font_header
    ws_dash["A11"].fill = fill_slate
    ws_dash["A11"].border = border_grid
    
    ws_dash["B11"] = "Count"
    ws_dash["B11"].font = font_header
    ws_dash["B11"].fill = fill_slate
    ws_dash["B11"].alignment = align_right
    ws_dash["B11"].border = border_grid
    
    audit_metrics = [
        ("Raw CRM Records Scanned", stats["initial_rows"]),
        ("Duplicate Leads Removed", stats["duplicates_removed"]),
        ("Missing Contact Emails Fixed", stats["missing_emails_filled"]),
        ("Lead Source Casing Normalized", stats["casing_fixes"]),
        ("Out-of-Order Dates Corrected", stats["date_order_fixes"]),
        ("MRR Valuation Anomalies Cleaned", stats["mrr_fixes"]),
        ("Final Cleaned CRM Records", stats["final_clean_rows"])
    ]
    
    for r_idx, (metric, count) in enumerate(audit_metrics, start=12):
        cell_a = ws_dash.cell(row=r_idx, column=1, value=metric)
        cell_b = ws_dash.cell(row=r_idx, column=2, value=count)
        
        cell_a.font = font_data
        cell_b.font = font_data_bold
        cell_a.border = border_grid
        cell_b.border = border_grid
        cell_b.alignment = align_right
        cell_b.number_format = "#,##0"
        
        if r_idx == 18: # Final Cleaned rows
            cell_a.fill = fill_zebra
            cell_b.fill = fill_zebra
            
    # -------------------------------------------------------------
    # TAB 2: WEEKLY METRICS (FORMULA-DRIVEN)
    # -------------------------------------------------------------
    ws_week = wb.create_sheet(title="Weekly Metrics")
    ws_week.views.sheetView[0].showGridLines = True
    
    # Title
    ws_week.merge_cells("A1:H2")
    ws_week_title = ws_week["A1"]
    ws_week_title.value = " Weekly Funnel & Revenue Tracking Table"
    ws_week_title.font = font_title
    ws_week_title.fill = fill_navy
    ws_week_title.alignment = Alignment(horizontal='left', vertical='center')
    
    # Table Headers
    headers = [
        "Week Start Date", 
        "Leads Generated", 
        "Trials Activated", 
        "Conversions (Closed Won)", 
        "Lead-to-Trial Rate", 
        "Trial-to-Conversion Rate", 
        "Weekly MRR Added",
        "Avg Velocity (Lead->Conv)"
    ]
    
    for c_idx, header in enumerate(headers, start=1):
        cell = ws_week.cell(row=4, column=c_idx, value=header)
        cell.font = font_header
        cell.fill = fill_slate
        cell.alignment = align_center
        cell.border = border_grid
        
    # Write week start dates (13 weeks starting 2026-03-01 to 2026-05-24)
    week_starts = [
        date(2026, 3, 1) + timedelta(weeks=w) for w in range(13)
    ]
    
    for r_offset, w_start in enumerate(week_starts):
        row = 5 + r_offset
        
        # Column A: Week Start Date
        c_a = ws_week.cell(row=row, column=1, value=w_start)
        c_a.number_format = 'yyyy-mm-dd'
        c_a.alignment = align_center
        
        # Column B: Leads (COUNTIFS in Cleaned Data Lead_Date)
        # B6:B500+ is Lead_Date
        # In the formula, we reference 'Cleaned Data'!$B$2:$B$600
        # Formula counts rows where Lead_Date >= current week start AND Lead_Date < next week start
        c_b = ws_week.cell(row=row, column=2)
        c_b.value = f"=COUNTIFS('Cleaned Data'!$B$2:$B$600, \">=\" & A{row}, 'Cleaned Data'!$B$2:$B$600, \"<\" & (A{row} + 7))"
        c_b.alignment = align_right
        
        # Column C: Trials
        c_c = ws_week.cell(row=row, column=3)
        c_c.value = f"=COUNTIFS('Cleaned Data'!$H$2:$H$600, \">=\" & A{row}, 'Cleaned Data'!$H$2:$H$600, \"<\" & (A{row} + 7), 'Cleaned Data'!$G$2:$G$600, \"Yes\")"
        c_c.alignment = align_right
        
        # Column D: Conversions
        c_d = ws_week.cell(row=row, column=4)
        c_d.value = f"=COUNTIFS('Cleaned Data'!$J$2:$J$600, \">=\" & A{row}, 'Cleaned Data'!$J$2:$J$600, \"<\" & (A{row} + 7), 'Cleaned Data'!$I$2:$I$600, \"Yes\")"
        c_d.alignment = align_right
        
        # Column E: Lead-to-Trial Rate = Trials / Leads
        c_e = ws_week.cell(row=row, column=5)
        c_e.value = f"=IF(B{row}=0, 0, C{row}/B{row})"
        c_e.number_format = '0.0%'
        c_e.alignment = align_right
        
        # Column F: Trial-to-Conversion Rate = Conversions / Trials
        c_f = ws_week.cell(row=row, column=6)
        c_f.value = f"=IF(C{row}=0, 0, D{row}/C{row})"
        c_f.number_format = '0.0%'
        c_f.alignment = align_right
        
        # Column G: Weekly MRR Added = SUMIFS in Cleaned Data MRR
        c_g = ws_week.cell(row=row, column=7)
        c_g.value = f"=SUMIFS('Cleaned Data'!$K$2:$K$600, 'Cleaned Data'!$J$2:$J$600, \">=\" & A{row}, 'Cleaned Data'!$J$2:$J$600, \"<\" & (A{row} + 7))"
        c_g.number_format = '$#,##0'
        c_g.alignment = align_right
        
        # Column H: Avg Velocity (Lead->Conv) = AVERAGEIFS of Total_Deal_Velocity (Col N) in Cleaned Data
        c_h = ws_week.cell(row=row, column=8)
        c_h.value = f"=IFERROR(AVERAGEIFS('Cleaned Data'!$N$2:$N$600, 'Cleaned Data'!$J$2:$J$600, \">=\" & A{row}, 'Cleaned Data'!$J$2:$J$600, \"<\" & (A{row} + 7)), \"-\")"
        c_h.alignment = align_center
        
        # Fonts, Borders and Zebra Striping
        for col in range(1, 9):
            cell = ws_week.cell(row=row, column=col)
            cell.font = font_data
            cell.border = border_grid
            if row % 2 == 0:
                cell.fill = fill_zebra
                
    # Add Total/Average Row
    tot_row = 5 + len(week_starts)
    
    ws_week.cell(row=tot_row, column=1, value="Total").font = font_data_bold
    ws_week.cell(row=tot_row, column=1).alignment = align_center
    ws_week.cell(row=tot_row, column=1).fill = fill_total
    ws_week.cell(row=tot_row, column=1).border = border_total
    
    # Leads Total
    c_b_tot = ws_week.cell(row=tot_row, column=2, value=f"=SUM(B5:B{tot_row-1})")
    c_b_tot.font = font_data_bold
    c_b_tot.alignment = align_right
    c_b_tot.fill = fill_total
    c_b_tot.border = border_total
    
    # Trials Total
    c_c_tot = ws_week.cell(row=tot_row, column=3, value=f"=SUM(C5:C{tot_row-1})")
    c_c_tot.font = font_data_bold
    c_c_tot.alignment = align_right
    c_c_tot.fill = fill_total
    c_c_tot.border = border_total
    
    # Conversions Total
    c_d_tot = ws_week.cell(row=tot_row, column=4, value=f"=SUM(D5:D{tot_row-1})")
    c_d_tot.font = font_data_bold
    c_d_tot.alignment = align_right
    c_d_tot.fill = fill_total
    c_d_tot.border = border_total
    
    # Weighted Lead-to-Trial Rate
    c_e_tot = ws_week.cell(row=tot_row, column=5, value=f"=C{tot_row}/B{tot_row}")
    c_e_tot.font = font_data_bold
    c_e_tot.number_format = '0.0%'
    c_e_tot.alignment = align_right
    c_e_tot.fill = fill_total
    c_e_tot.border = border_total
    
    # Weighted Trial-to-Conversion Rate
    c_f_tot = ws_week.cell(row=tot_row, column=6, value=f"=D{tot_row}/C{tot_row}")
    c_f_tot.font = font_data_bold
    c_f_tot.number_format = '0.0%'
    c_f_tot.alignment = align_right
    c_f_tot.fill = fill_total
    c_f_tot.border = border_total
    
    # Total MRR Added
    c_g_tot = ws_week.cell(row=tot_row, column=7, value=f"=SUM(G5:G{tot_row-1})")
    c_g_tot.font = font_data_bold
    c_g_tot.number_format = '$#,##0'
    c_g_tot.alignment = align_right
    c_g_tot.fill = fill_total
    c_g_tot.border = border_total
    
    # Average overall Deal Velocity
    c_h_tot = ws_week.cell(row=tot_row, column=8, value=f"=AVERAGE('Cleaned Data'!$N$2:$N$600)")
    c_h_tot.font = font_data_bold
    c_h_tot.alignment = align_center
    c_h_tot.fill = fill_total
    c_h_tot.border = border_total
    c_h_tot.number_format = '0.0'
    
    # -------------------------------------------------------------
    # TAB 3: CLEANED DATA
    # -------------------------------------------------------------
    ws_data = wb.create_sheet(title="Cleaned Data")
    ws_data.views.sheetView[0].showGridLines = True
    
    # Headers
    data_headers = list(df.columns) + ["Lead_to_Trial_Days", "Trial_to_Conversion_Days", "Total_Deal_Velocity"]
    
    for c_idx, d_hdr in enumerate(data_headers, start=1):
        cell = ws_data.cell(row=1, column=c_idx, value=d_hdr)
        cell.font = font_header
        cell.fill = fill_slate
        cell.alignment = align_center
        cell.border = border_grid
        
    # Write Cleaned Data Rows
    for r_idx, row in df.iterrows():
        excel_row = r_idx + 2 # index starts at 0, Excel starts at 2
        
        # Write values
        ws_data.cell(row=excel_row, column=1, value=row["Lead_ID"]).alignment = align_center
        
        c_ld = ws_data.cell(row=excel_row, column=2, value=row["Lead_Date"])
        c_ld.number_format = 'yyyy-mm-dd'
        c_ld.alignment = align_center
        
        ws_data.cell(row=excel_row, column=3, value=row["Company_Name"]).alignment = align_left
        ws_data.cell(row=excel_row, column=4, value=row["Contact_Email"]).alignment = align_left
        ws_data.cell(row=excel_row, column=5, value=row["Lead_Source"]).alignment = align_center
        ws_data.cell(row=excel_row, column=6, value=row["Segment"]).alignment = align_center
        ws_data.cell(row=excel_row, column=7, value=row["Trial_Activated"]).alignment = align_center
        
        c_td = ws_data.cell(row=excel_row, column=8, value=row["Trial_Date"])
        c_td.number_format = 'yyyy-mm-dd'
        c_td.alignment = align_center
        
        ws_data.cell(row=excel_row, column=9, value=row["Converted"]).alignment = align_center
        
        c_cd = ws_data.cell(row=excel_row, column=10, value=row["Conversion_Date"])
        c_cd.number_format = 'yyyy-mm-dd'
        c_cd.alignment = align_center
        
        # MRR
        c_m = ws_data.cell(row=excel_row, column=11, value=row["MRR"])
        c_m.number_format = '$#,##0'
        c_m.alignment = align_right
        
        # Column 12: Lead_to_Trial_Days (formula: =IF(G2="Yes", H2-B2, ""))
        c_ltd = ws_data.cell(row=excel_row, column=12, value=f"=IF(G{excel_row}=\"Yes\", H{excel_row}-B{excel_row}, \"\")")
        c_ltd.alignment = align_right
        
        # Column 13: Trial_to_Conversion_Days (formula: =IF(I2="Yes", J2-H2, ""))
        c_tcd = ws_data.cell(row=excel_row, column=13, value=f"=IF(I{excel_row}=\"Yes\", J{excel_row}-H{excel_row}, \"\")")
        c_tcd.alignment = align_right
        
        # Column 14: Total_Deal_Velocity (formula: =IF(I2="Yes", J{excel_row}-B{excel_row}, ""))
        c_tdv = ws_data.cell(row=excel_row, column=14, value=f"=IF(I{excel_row}=\"Yes\", J{excel_row}-B{excel_row}, \"\")")
        c_tdv.alignment = align_right
        
        # Fonts, Borders and Zebra Striping
        for col in range(1, 15):
            cell = ws_data.cell(row=excel_row, column=col)
            cell.font = font_data
            cell.border = border_grid
            if excel_row % 2 == 0:
                cell.fill = fill_zebra
                
    # -------------------------------------------------------------
    # CHARTS IN EXECUTIVE DASHBOARD
    # -------------------------------------------------------------
    # Add Weekly Conversion Line Chart to Dashboard
    chart_funnel = LineChart()
    chart_funnel.title = "Weekly Marketing & Sales Funnel Trends"
    chart_funnel.style = 13
    chart_funnel.y_axis.title = "Count"
    chart_funnel.x_axis.title = "Week Start Date"
    chart_funnel.width = 16
    chart_funnel.height = 10
    
    # Reference Leads, Trials, Conversions columns in Weekly Metrics
    # Sheet: 'Weekly Metrics', range B4:D17 (Headers + 13 weeks)
    data_ref = Reference(ws_week, min_col=2, min_row=4, max_col=4, max_row=17)
    cats_ref = Reference(ws_week, min_col=1, min_row=5, max_row=17)
    
    chart_funnel.add_data(data_ref, titles_from_data=True)
    chart_funnel.set_categories(cats_ref)
    
    # Place chart in Dashboard tab
    ws_dash.add_chart(chart_funnel, "D11")
    
    # Add Weekly MRR Column Chart
    chart_mrr = BarChart()
    chart_mrr.type = "col"
    chart_mrr.style = 10
    chart_mrr.title = "Weekly New MRR Added"
    chart_mrr.y_axis.title = "MRR ($)"
    chart_mrr.x_axis.title = "Week Start Date"
    chart_mrr.width = 16
    chart_mrr.height = 10
    chart_mrr.legend = None # No legend needed as it's a single series
    
    mrr_ref = Reference(ws_week, min_col=7, min_row=4, max_row=17)
    chart_mrr.add_data(mrr_ref, titles_from_data=True)
    chart_mrr.set_categories(cats_ref)
    
    # Place chart in Dashboard tab
    ws_dash.add_chart(chart_mrr, "D31")
    
    # -------------------------------------------------------------
    # AUTO-FIT COLUMN WIDTHS ACROSS ALL SHEETS
    # -------------------------------------------------------------
    for ws in [ws_dash, ws_week, ws_data]:
        for col in ws.columns:
            # Skip merged cells width calculations
            is_merged = False
            for cell in col:
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        is_merged = True
                        break
                if is_merged:
                    break
            
            # Basic auto-fit logic
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if val.startswith("="):
                    # Guess formula length to avoid narrow cols
                    max_len = max(max_len, 12)
                else:
                    max_len = max(max_len, len(val))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    # Explicit override for specific column widths to ensure perfect layout
    ws_dash.column_dimensions['A'].width = 32
    ws_dash.column_dimensions['B'].width = 12
    ws_dash.column_dimensions['C'].width = 12
    ws_dash.column_dimensions['D'].width = 14
    ws_dash.column_dimensions['E'].width = 14
    ws_dash.column_dimensions['F'].width = 14
    ws_dash.column_dimensions['G'].width = 14
    ws_dash.column_dimensions['H'].width = 14
    
    # Save Workbook
    filename = "reports/revenue_pipeline_tracker.xlsx"
    wb.save(filename)
    print(f"Successfully generated styled Excel tracker in '{filename}'!")

if __name__ == "__main__":
    df, stats = clean_and_process_data()
    create_excel_tracker(df, stats)
